from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from .modelos import Filamento
from .excel_funciones import *
from .historial import *
from .funciones_json_backup import cargar_stock

def exportar_stock_a_excel(stock_filamentos):
    workbook = Workbook()
    hoja = workbook.active
    hoja.title = "Stock"

    hoja.append(["Color", "Fabricante", "Cantidad"])

    for filamento in stock_filamentos:
        hoja.append([filamento.color, filamento.fabricante, filamento.cantidad])

    ultima_fila = hoja.max_row

    tabla = Table(displayName="TablaStock", ref=f"A1:C{ultima_fila}")

    estilo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    tabla.tableStyleInfo = estilo
    hoja.add_table(tabla)

    hoja.column_dimensions["A"].width = 20
    hoja.column_dimensions["B"].width = 18
    hoja.column_dimensions["C"].width = 12

    workbook.save("stock.xlsx")

    workbook.save("stock.xlsx")

def importar_stock_desde_excel():
    try:
        workbook = load_workbook("stock.xlsx")
        hoja = workbook["Stock"]

        stock_filamentos = []

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            color, fabricante, cantidad = fila

            if color is not None and fabricante is not None and cantidad is not None:
                # ambos por si maniana quiero hacer algo cuando la fila esta incompleta
                if color is None and fabricante is None and cantidad is None:
                    continue

                if color is None or fabricante is None or cantidad is None:
                    continue

                try:
                    cantidad = int(cantidad)
                except ValueError:
                    continue
                
                
                stock_filamentos.append(
                    Filamento(
                        str(color).strip().capitalize(),
                        str(fabricante).strip().capitalize(),
                        int(cantidad)
                    )
                )
        stock_filamentos.sort(key=lambda filamento: (filamento.color.lower(), filamento.fabricante.lower()))
        
        return stock_filamentos


    except FileNotFoundError:
        print("- No se encontró stock.xlsx, se cargará desde JSON -")
        return cargar_stock()